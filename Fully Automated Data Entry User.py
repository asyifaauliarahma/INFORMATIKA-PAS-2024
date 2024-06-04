import pathlib
from tkinter import *
from tkinter import messagebox
from tkinter.ttk import Combobox

import openpyxl
from openpyxl import Workbook

root = Tk()
root.title("Data Siswa Kelas 10 SMK TELKOM PURWOKERTO")
root.geometry('800x500+200+300')
root.resizable(False, False)
root.configure(bg="light blue")

file=pathlib.Path('Backened_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="NAMA"
    sheet['B1']="NIS"
    sheet['C1']="NO HP"
    sheet['D1']="TTL"
    sheet['E1']="KELAS"
    sheet['F1']="JENIS KELAMIN"
    sheet['G1']="ABSEN"
    sheet['H1'] = "EKSKUL"
    sheet['I1']="ALAMAT"


    file.save('Backened_data.xlsx')


def submit():
    name=nameValue.get()
    nis=nisValue.get()
    contact=contactValue.get()
    age=ttlValue.get()
    gender=gender_combobox.get()
    kelas=kelas_combobox.get()
    absen=absen_combobox.get()
    ekskul=ekskul_combobox.get()
    address=addressEntry.get(1.0, END)

    file=openpyxl.load_workbook('Backened_data.xlsx')
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=name)
    sheet.cell(column=2, row=sheet.max_row, value=nis)
    sheet.cell(column=3,row=sheet.max_row,value=contact)
    sheet.cell(column=4,row=sheet.max_row,value=age)
    sheet.cell(column=5, row=sheet.max_row, value=kelas)
    sheet.cell(column=6, row=sheet.max_row, value=gender)
    sheet.cell(column=7, row=sheet.max_row, value=absen)
    sheet.cell(column=8, row=sheet.max_row, value=ekskul)
    sheet.cell(column=9,row=sheet.max_row,value=address)



    file.save(r'Backened_data.xlsx')

    messagebox.showinfo('info', 'Berhasil Di Simpan!')

    nameValue.set('')
    nisValue.set('')
    contactValue.set('')
    ttlValue.set('')
    addressEntry.delete(1.0,END)


def clear():
    nameValue.set('')
    nisValue.set('')
    contactValue.set('')
    ttlValue.set('')
    addressEntry.delete(1.0,END)


# icon
root.iconbitmap(r'TELKOM.ico')

# heading
Label(root, text="Silahkan isi from di bawah ini >.< ", font="arial 13", bg= "light blue", fg="Black").place(x=20, y=20)

# Label

Label(root, text='NAMA', font=23, bg="light blue", fg="Black").place(x=70, y=105)
Label(root, text='NIS', font=23, bg="light blue", fg="Black").place(x=70, y=135)
Label(root, text='NO HP', font=23, bg="light blue", fg="Black").place(x=70, y=163)
Label(root, text='TTL', font=23, bg="light blue", fg="Black").place(x=70, y=200)
Label(root, text='JENIS ', font=23, bg="light blue", fg="Black").place(x=70, y=238)
Label(root, text='KELAS', font=23, bg="light blue", fg="Black").place(x=70, y=271)
Label(root, text='ABSEN', font=23, bg="light blue", fg="Black").place(x=360, y=235)
Label(root, text='EKSKUL', font=23, bg="light blue", fg="Black").place(x=360, y=270)
Label(root, text='ALAMAT', font=23, bg="light blue", fg="Black").place(x=70, y=310)


# Entry
nameValue = StringVar()
nisValue = StringVar()
contactValue = StringVar()
ttlValue = StringVar()


nameEntry = Entry(root, textvariable=nameValue, width=54, bd=2, font=20)
nisEntry = Entry(root, textvariable=nisValue, width=54, bd=2, font=20)
contactEntry = Entry(root, textvariable=contactValue,width=54, bd=2, font=20)
ttlEntry = Entry(root, textvariable=ttlValue, width=54, bd=2, font=20)


# gender
gender_combobox = Combobox(root, values=['Laki - Laki', 'Perempuan'], font='arial 12', state='r', width=14)
gender_combobox.place(x=200, y=235)


#kelas
kelas_combobox = Combobox(root, values=['X PPLG 1', 'X PPLG 2', 'X PPLG 3', 'X PPLG 4', 'X PPLG 5', 'X PPLG 6', 'X PPLG 7',
                                        'X TJKT 1', 'X TJKT 2', 'X TJKT 3', 'X TJKT 4', 'X TJKT 5',
                                        'XI RPL 1', 'XI RPL 2', 'XI RPL 3', 'XI RPL 4', 'XI RPL 5', 'XI RPL 6', 'XI RPL 7',
                                        'XI TJAT 1', 'XI TJAT 2','XI TKJ 1', 'XI TKJ 2', 'XI TKJ 3',
                                        'XII RPL 1', 'XII RPL 2', 'XII RPL 3', 'XII RPL 4', 'XII RPL 5',
                                        'XII TJA 1', 'XII TJA 2', 'XII TKJ 1', 'XII TKJ 2', 'XII TKJ 3', 'XII TKJ 4'],
                                        font='arial 12', state='r', width=14)
kelas_combobox.place(x=200, y=270)

#Absen
absen_combobox = Combobox(root, values=['1', '2', '3', '4', '5', '6', '7', '8', '9', '10',
                                        '11', '12', '13', '14', '15', '16', '17','18','19','20',
                                        '21', '22', '23', '24', '25', '26', '27', '28', '29', '30',
                                        '31', '32', '33', '34', '35',], font='arial 10', state='r', width=33)
absen_combobox.place(x=440, y=233)


#Ekskul
ekskul_combobox = Combobox(root, values=['PRAMUKA', 'PMR', 'TARI', 'PADUS', 'PASKIB', 'WEBTECH', 'VOLI',
                                         'BASKET', 'BADMINTON', 'IT SOFTWARE', 'ROBOTIK', 'DESAIN GRAFIS',
                                         'FUTSAL', 'HADROH', 'CYBER SECURITY', 'E-SPORT', 'PHOTOGRAFI',
                                         'MATEMATIK','MUSIK','BAHASA INGGRIS','BELA DIRI', 'JURNALISTIK',
                                         'PKS', 'AI ARTIFICIAL INTELLIGENCE', 'ENTREPENEUR', 'ANIMASI',
                                         'INFORMATION NETWORK CABLING',], font='arial 10', state='r', width=33)
ekskul_combobox.place(x=440, y=270)






addressEntry = Text(root, width=50, height=4, bd=2)


nameEntry.place(x=200, y=105)
nisEntry.place(x=200, y=136)
contactEntry.place(x=200, y=167)
ttlEntry.place(x=200, y=202)
addressEntry.place(x=200, y=315)


Button(root, text="Kirim", bg="Light Blue", fg="Black", width=15, height=2, command=submit).place(x=200, y=410)
Button(root, text="Hapus", bg="Light Blue", fg="Black", width=15, height=2, command=clear).place(x=340, y=410)
Button(root, text="Keluar", bg="Light Blue", fg="Black", width=15, height=2, command=lambda: root.destroy()).place(x=480,
y=410)

root.mainloop()